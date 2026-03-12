#-------------------------------------------------------------------
# YASKAWAapp_bckend/views.py
#-------------------------------------------------------------------
# 1. Standards Python
import re
import io
import base64
import pickle
from datetime import datetime

# 2. Bibliothèques tierces
import openpyxl
import pandas as pd

# 3. Django : Core, Auth & Models
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils.text import slugify

# 4. Imports Locaux (l'application)
from .models import Category, Product, Compatibility, ProductSpec
from .filters import ProductFilter, ServopackFilter, MotorFilter
from .parsers.yaskawa_parser import YaskawaParser


# ── Helper ─────────────────────────────────────────────────────────────────
def is_admin_or_engineer(user):
    return user.is_staff or user.groups.filter(name='Engineer').exists()

#-------------------------------------------------------------------
# 1. Connexion
#-------------------------------------------------------------------
def login_view(request):
    if request.method == 'POST':
        u    = request.POST.get('username')
        p    = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            if user.is_active:
                login(request, user)
                return redirect('/admin/') if user.is_superuser else redirect('home')
            return render(request, 'auth/login.html', {'error': "Compte en attente d'activation par un administrateur."})
        return render(request, 'auth/login.html', {'error': "Identifiants invalides."})
    return render(request, 'auth/login.html')

#-------------------------------------------------------------------
# 2. Inscription
#-------------------------------------------------------------------
def register_view(request):
    if request.method == 'POST':
        email       = request.POST.get('username')
        pwd         = request.POST.get('password')
        cpwd        = request.POST.get('confirm_password')
        service_nom = request.POST.get('service')

        if not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email):
            return render(request, 'auth/register.html', {'error': "Format d'e-mail invalide."})
        reg = r"^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&#.])[A-Za-z\d@$!%*?&#.]{12,}$"
        if not re.match(reg, pwd):
            return render(request, 'auth/register.html', {'error': "MDP : 12 car. min, 1 Maj, 1 Min, 1 Chiffre, 1 Spécial."})
        if pwd != cpwd:
            return render(request, 'auth/register.html', {'error': "Les mots de passe ne correspondent pas."})
        if User.objects.filter(username=email).exists():
            return render(request, 'auth/register.html', {'error': "Cet e-mail est déjà utilisé."})

        user = User.objects.create_user(username=email, email=email, password=pwd)
        user.is_active = False
        user.save()
        if service_nom:
            group, _ = Group.objects.get_or_create(name=service_nom)
            user.groups.add(group)
        return render(request, 'auth/login.html', {'error': "Compte créé ! Attendez l'activation admin."})
    return render(request, 'auth/register.html')

#-------------------------------------------------------------------
# 3. Home
#-------------------------------------------------------------------
@login_required
def home_view(request):
    is_engineer = request.user.groups.filter(name='Engineer').exists()
    return render(request, 'home.html', {
        'categories':  Category.objects.all(),
        'is_engineer': is_engineer,
        'can_update':  request.user.is_staff or is_engineer,
    })

#-------------------------------------------------------------------
# 4. Liste produits d'une catégorie
#-------------------------------------------------------------------
#-------------------------------------------------------------------
# 4. Liste produits d'une catégorie (MODIFIÉ)
#-------------------------------------------------------------------
@login_required
def category_detail_view(request, slug):
    category    = get_object_or_404(Category, slug=slug)
    is_engineer = request.user.groups.filter(name='Engineer').exists()
    all_products_qs = category.products.all()

    sgd_qs = all_products_qs.filter(Q(model_code__icontains='SGD') | Q(product_type__icontains='Servopack'))
    sgm_qs = all_products_qs.filter(Q(model_code__icontains='SGM') | Q(product_type__icontains='Motor'))

    clean_params = {k: v for k, v in request.GET.items() if v != ''}

    sgd_filter   = ServopackFilter(clean_params, queryset=sgd_qs)
    sgm_filter   = MotorFilter(clean_params, queryset=sgm_qs)

    sgd_page_obj = Paginator(sgd_filter.qs, 8).get_page(request.GET.get('sgd_page'))
    sgm_page_obj = Paginator(sgm_filter.qs, 8).get_page(request.GET.get('sgm_page'))

    # --- PRÉPARATION DES FILTRES NETTOYÉS ---
    
    # 1. Voltages (Exclure vide et None)
    voltages = all_products_qs.filter(spec__voltage__isnull=False) \
               .exclude(spec__voltage__in=['', 'None']) \
               .values_list('spec__voltage', flat=True).distinct().order_by('spec__voltage')

    # 2. Puissances
    power_values = [
        str(p).replace(',', '.')
        for p in all_products_qs.filter(spec__power_kw__isnull=False)
        .values_list('spec__power_kw', flat=True).distinct().order_by('spec__power_kw')
        if p is not None
    ]

    # 3. Protocoles SGD (Exclure vide, None et 'Other')
    sgd_protocols = sgd_qs.values_list('spec__protocol', flat=True).distinct() \
                    .exclude(spec__protocol__in=['', 'None', 'Other']) \
                    .order_by('spec__protocol')

    # 4. Encodeurs SGM (Exclure vide, None et 'Other')
    sgm_encoders = sgm_qs.values_list('spec__encoder_type', flat=True).distinct() \
                   .exclude(spec__encoder_type__in=['', 'None', 'Other']) \
                   .order_by('spec__encoder_type')

    # 5. Séries
    sgd_series = sgd_qs.values_list('spec__series', flat=True).distinct().exclude(spec__series__in=['', 'None'])
    sgm_series = sgm_qs.values_list('spec__series', flat=True).distinct().exclude(spec__series__in=['', 'None'])

    return render(request, 'products_list.html', {
        'category':      category,
        'is_engineer':   is_engineer,
        'sgd_products':  sgd_page_obj,
        'sgm_products':  sgm_page_obj,
        'voltages':      voltages,
        'power_values':  power_values,
        'sgd_protocols': sgd_protocols,
        'sgd_series':    sgd_series,
        'sgm_series':    sgm_series,
        'sgm_encoders':  sgm_encoders,
        'search_query':  request.GET.get('search', ''),
    })
#-------------------------------------------------------------------
# 5. Détail produit
#-------------------------------------------------------------------
@login_required
def product_detail_view(request, model_code):
    product = get_object_or_404(Product, model_code=model_code)
    return render(request, 'product_detail.html', {'product': product})

#-------------------------------------------------------------------
# 6. Page intermédiaire : choisir le produit manquant (pick_pair)
#
#  Appelée depuis :
#    - product_detail  → "Continuer ma solution"  (1 produit connu)
#    - product_list    → "Build my solution"       (session contient sp + mt)
#
#  Logique :
#    - Si model_code est un SGD  → on propose les moteurs compatibles
#    - Si model_code est un SGM  → on propose les servopacks compatibles
#-------------------------------------------------------------------
@login_required
def pick_pair_view(request, model_code):
    known_product = get_object_or_404(Product, model_code=model_code)
    ref           = known_product.model_code.upper()

    # Récupère les specs du produit connu
    try:
        known_spec = known_product.spec
    except:
        known_spec = None

    if ref.startswith('SGD'):
        looking_for = 'motor'
        pair_label  = 'Motor'
        sp_code     = model_code

        candidates = Product.objects.filter(
            Q(model_code__icontains='SGM') | Q(product_type__icontains='Motor')
        )
        # Filtre par voltage et puissance si les specs existent
        if known_spec:
            if known_spec.voltage:
                candidates = candidates.filter(spec__voltage=known_spec.voltage)
            if known_spec.power_kw:
                candidates = candidates.filter(spec__power_kw=known_spec.power_kw)

    else:
        looking_for = 'servopack'
        pair_label  = 'Servopack'
        sp_code     = None

        candidates = Product.objects.filter(
            Q(model_code__icontains='SGD') | Q(product_type__icontains='Servopack')
        )
        if known_spec:
            if known_spec.voltage:
                candidates = candidates.filter(spec__voltage=known_spec.voltage)
            if known_spec.power_kw:
                candidates = candidates.filter(spec__power_kw=known_spec.power_kw)

    candidates = candidates.distinct()

    # Fallback : si aucun résultat avec les specs, affiche tout le type
    if not candidates.exists():
        if ref.startswith('SGD'):
            candidates = Product.objects.filter(
                Q(model_code__icontains='SGM') | Q(product_type__icontains='Motor')
            ).distinct()
        else:
            candidates = Product.objects.filter(
                Q(model_code__icontains='SGD') | Q(product_type__icontains='Servopack')
            ).distinct()

    page_obj = Paginator(candidates, 12).get_page(request.GET.get('page'))

    return render(request, 'pick_pair.html', {
        'known_product': known_product,
        'pair_label':    pair_label,
        'looking_for':   looking_for,
        'candidates':    page_obj,
        'sp_code':       sp_code,
    })
#-------------------------------------------------------------------
# 7. Page résumé : valider la paire avant de lancer le configurateur
#-------------------------------------------------------------------
@login_required
def resume_view(request, sp_code, mt_code):
    servopack = get_object_or_404(Product, model_code=sp_code)
    motor     = get_object_or_404(Product, model_code=mt_code)

    # Sauvegarde la paire en session pour la retrouver dans my_solution
    request.session['selected_servopack'] = sp_code
    request.session['selected_motor']     = mt_code
    request.session.modified              = True

    return render(request, 'product_resume.html', {
        'servopack': servopack,
        'motor':     motor,
        'sp_code':   sp_code,
        'mt_code':   mt_code,
    })

#-------------------------------------------------------------------
# 8. Vues de sélection session (boutons "Select" sur product_list)
#-------------------------------------------------------------------
@login_required
def select_servopack_view(request, sp_code):
    if sp_code == 'none':
        request.session.pop('selected_servopack', None)
    else:
        request.session['selected_servopack'] = sp_code
    return redirect(request.META.get('HTTP_REFERER', 'home'))

@login_required
def select_motor_view(request, m_code):
    if m_code == 'none':
        request.session.pop('selected_motor', None)
    else:
        request.session['selected_motor'] = m_code
    return redirect(request.META.get('HTTP_REFERER', 'home'))

#-------------------------------------------------------------------
# 9. Import Base de Données
#-------------------------------------------------------------------
@login_required
def import_data_view(request):
    categories   = Category.objects.all()
    base_context = {'step': 'upload', 'categories': categories}

    if request.method == 'POST':
        category_id = request.POST.get('category_id')
        import_type = request.POST.get('import_type', 'products')

        # ── CONFIRMATION ──────────────────────────────────────────────────
        if 'confirm' in request.POST:
            file_data   = request.session.get('import_file_data')
            category_id = request.session.get('import_category_id')
            import_type = request.session.get('import_type', 'products')
            col_code    = request.session.get('import_col_code', 'ModelCode')

            if not file_data:
                base_context['error'] = "Session expirée, veuillez ré-uploader le fichier."
                return render(request, 'import_data.html', base_context)

            df = pickle.loads(base64.b64decode(file_data))
            added, updated, errors = 0, 0, []

            # ── Products ──────────────────────────────────────────────────
            if import_type == 'products':
                forced_category = Category.objects.filter(id=category_id).first() if category_id else None
                for index, row in df.iterrows():
                    try:
                        m_code = str(row.get('ModelCode', '')).strip()
                        if not m_code or m_code == 'nan':
                            continue
                        if forced_category:
                            category_obj = forced_category
                        else:
                            c_name = str(row.get('Category', 'Non classé')).strip()
                            category_obj, _ = Category.objects.get_or_create(
                                name=c_name, defaults={'slug': slugify(c_name)}
                            )
                        product_obj, created = Product.objects.update_or_create(
                            model_code=m_code,
                            defaults={
                                'category':        category_obj,
                                'numero_SAP':      str(row.get('SAPNo',       '') or '').strip(),
                                'product_type':    str(row.get('Product',     '') or '').strip(),
                                'features':        str(row.get('Attribute1',  '') or '').strip(),
                                'specification':   str(row.get('Attribute2',  '') or '').strip(),
                                'compatible':      str(row.get('Attribute3',  '') or '').strip(),
                                'usable_with':     str(row.get('Attribute4',  '') or '').strip(),
                                'compat_range_1':  str(row.get('Attribute5',  '') or '').strip(),
                                'compat_range_2':  str(row.get('Attribute6',  '') or '').strip(),
                                'compat_range_3':  str(row.get('Attribute7',  '') or '').strip(),
                                'compat_range_4':  str(row.get('Attribute8',  '') or '').strip(),
                                'compat_range_5':  str(row.get('Attribute9',  '') or '').strip(),
                                'compat_range_6':  str(row.get('Attribute10', '') or '').strip(),
                                'compat_range_7':  str(row.get('Attribute11', '') or '').strip(),
                                'compat_range_8':  str(row.get('Attribute12', '') or '').strip(),
                                'compat_range_9':  str(row.get('Attribute13', '') or '').strip(),
                                'compat_range_10': str(row.get('Attribute14', '') or '').strip(),
                                'compat_range_11': str(row.get('Attribute15', '') or '').strip(),
                            }
                        )
                        if created: added += 1
                        else:       updated += 1
                        parsed = YaskawaParser.parse(m_code)
                        if parsed:
                            ProductSpec.objects.update_or_create(
                                product=product_obj,
                                defaults={
                                    'series':        parsed.get('series',        ''),
                                    'power_kw':      parsed.get('power_kw',      None),
                                    'voltage':       parsed.get('voltage',       ''),
                                    'protocol':      parsed.get('protocol',      ''),
                                    'encoder_type':  parsed.get('encoder_type',  ''),
                                    'option':        parsed.get('option',        ''),
                                    'specification': parsed.get('specification', ''),
                                    'brake':         parsed.get('brake',         False),
                                }
                            )
                    except Exception as e:
                        errors.append(f"Ligne {index + 2} : {e}")

            # ── Compatibilities ───────────────────────────────────────────
            elif import_type == 'compatibilities':
                product_cache  = {p.model_code: p for p in Product.objects.all()}
                existing_links = set(Compatibility.objects.values_list(
                    'source_product__model_code', 'compatible_product__model_code'
                ))
                new_links = []
                for index, row in df.iterrows():
                    try:
                        m1 = str(row.get('ModelCode1', '')).strip()
                        m2 = str(row.get('ModelCode2', '')).strip()
                        p1, p2 = product_cache.get(m1), product_cache.get(m2)
                        if p1 and p2:
                            if (m1, m2) not in existing_links:
                                new_links.append(Compatibility(
                                    source_product=p1,
                                    compatible_product=p2,
                                    relation_type='Compatible'
                                ))
                                existing_links.add((m1, m2))
                                added += 1
                            else:
                                updated += 1
                        else:
                            errors.append(f"Ligne {index + 2} : '{m1}' ou '{m2}' introuvable.")
                    except Exception as e:
                        errors.append(f"Ligne {index + 2} : {e}")
                if new_links:
                    Compatibility.objects.bulk_create(new_links, batch_size=1000)

            # ── Stock ─────────────────────────────────────────────────────
            elif import_type == 'stock':
                stock_in, stock_out, stock_unknown = _process_stock(df, col_code, write=True)
                request.session.pop('import_file_data', None)
                return render(request, 'import_data.html', {
                    'step':          'result',
                    'categories':    categories,
                    'import_type':   'stock',
                    'stock_in':      stock_in,
                    'stock_out':     stock_out,
                    'stock_unknown': stock_unknown,
                })

            request.session.pop('import_file_data', None)
            return render(request, 'import_data.html', {
                'step':        'result',
                'categories':  categories,
                'import_type': import_type,
                'added':       added,
                'updated':     updated,
                'errors':      errors[:50],
            })

        # ── PREVIEW (upload fichier) ───────────────────────────────────────
        fichier = request.FILES.get('excel_file')
        if not fichier:
            base_context['error'] = "Aucun fichier sélectionné."
            return render(request, 'import_data.html', base_context)

        try:
            df         = pd.read_excel(fichier)
            df.columns = [str(c).strip() for c in df.columns]
            request.session['import_file_data']   = base64.b64encode(pickle.dumps(df)).decode()
            request.session['import_category_id'] = category_id
            request.session['import_type']        = import_type
        except Exception as e:
            base_context['error'] = f"Impossible de lire le fichier : {e}"
            return render(request, 'import_data.html', base_context)

        # Preview stock
        if import_type == 'stock':
            col_code = request.POST.get('col_code', 'ModelCode')
            request.session['import_col_code'] = col_code

            if col_code not in df.columns:
                base_context['error'] = (
                    f"Colonne '{col_code}' introuvable. "
                    f"Colonnes disponibles : {list(df.columns)}"
                )
                return render(request, 'import_data.html', base_context)

            stock_in, stock_out, stock_unknown = _process_stock(df, col_code, write=False)
            return render(request, 'import_data.html', {
                'step':          'preview',
                'categories':    categories,
                'import_type':   'stock',
                'col_code':      col_code,
                'total_rows':    len(df),
                'stock_in':      stock_in,
                'stock_out':     stock_out,
                'stock_unknown': stock_unknown,
            })

        # Preview produits / compatibilités
        return render(request, 'import_data.html', {
            'step':        'preview',
            'categories':  categories,
            'import_type': import_type,
            'headers':     list(df.columns),
            'preview':     df.head(10).values.tolist(),
            'total_rows':  len(df),
        })

    return render(request, 'import_data.html', base_context)


# ── Helper stock ───────────────────────────────────────────────────────────
def _process_stock(df, col_code, write=False):
    """
    Calcule (et optionnellement applique) les changements de stock.
    Retourne (nb_in_stock, nb_out_stock, nb_unknown).
    """
    stock_codes = set()
    for val in df[col_code].dropna():
        code = str(val).strip()
        if code and code != 'nan':
            stock_codes.add(code)

    all_products  = list(Product.objects.only('id', 'model_code', 'numero_SAP', 'in_stock'))
    by_model_code = {p.model_code: p for p in all_products}
    by_sap        = {p.numero_SAP: p for p in all_products if p.numero_SAP}

    to_in, to_out, unknown = [], [], []

    for code in stock_codes:
        product = by_model_code.get(code) or by_sap.get(code)
        if product:
            if not product.in_stock:
                product.in_stock = True
                to_in.append(product)
        else:
            unknown.append(code)

    for p in all_products:
        in_excel = (p.model_code in stock_codes) or (p.numero_SAP in stock_codes)
        if not in_excel and p.in_stock:
            p.in_stock = False
            to_out.append(p)

    if write:
        if to_in:
            Product.objects.bulk_update(to_in,  ['in_stock'], batch_size=500)
        if to_out:
            Product.objects.bulk_update(to_out, ['in_stock'], batch_size=500)

    return len(to_in), len(to_out), len(unknown)

#-------------------------------------------------------------------
# 11. Configurateur (BOM)
#-------------------------------------------------------------------
@login_required
def my_solution_view(request, sp_code, mt_code=None):
    servopack     = get_object_or_404(Product, model_code=sp_code)
    motor_product = Product.objects.filter(model_code=mt_code).first() if mt_code else None

    # Gestion BOM session
    current_source = request.session.get('current_source_code')
    if current_source != sp_code:
        bom_list = []
        request.session['current_source_code'] = sp_code
    else:
        bom_list = list(request.session.get('bom_list', []))

    request.session['bom_list'] = bom_list
    request.session.modified    = True

    # Produits protégés — jamais dans les onglets ni supprimables
    protected = {sp_code}
    if motor_product:
        protected.add(motor_product.model_code)

    def classify(comp):
        ref   = comp.compatible_product.model_code.upper()
        ptype = (getattr(comp.compatible_product, 'product_type', '') or '').lower()
        if ref.startswith('JZSP-CS') or ref.startswith('JZSP-S') or 'encoder' in ptype:
            return 'encoder_cable'
        if ref.startswith('JZSP-M') or ref.startswith('JZSP-U') or 'CBK' in ref or 'power' in ptype:
            return 'power_cable'
        if ref.startswith('SGM') or ref.startswith('SGD'):
            return 'skip'
        return 'accessory'

    # Compatibilités depuis le moteur (encoder + power cables)
    tab1, tab2, tab3 = [], [], []

    if motor_product:
        motor_compat = Compatibility.objects.filter(
            source_product=motor_product
        ).select_related('compatible_product').exclude(
            compatible_product__model_code__in=list(protected) + bom_list
        )
        for comp in motor_compat:
            cat = classify(comp)
            if   cat == 'encoder_cable': tab1.append(comp)
            elif cat == 'power_cable':   tab2.append(comp)
            elif cat == 'accessory':     tab3.append(comp)

    # Compatibilités depuis le servopack (accessories uniquement)
    sgd_compat = Compatibility.objects.filter(
        source_product=servopack
    ).select_related('compatible_product').exclude(
        compatible_product__model_code__in=list(protected) + bom_list
    )
    seen = {c.compatible_product.model_code for c in tab1 + tab2 + tab3}
    for comp in sgd_compat:
        cat  = classify(comp)
        code = comp.compatible_product.model_code
        if code in seen or cat == 'skip':
            continue
        seen.add(code)
        if   cat == 'encoder_cable': tab1.append(comp)
        elif cat == 'power_cable':   tab2.append(comp)
        elif cat == 'accessory':     tab3.append(comp)

    return render(request, 'my_solution.html', {
        'main_product':  servopack,
        'motor_product': motor_product,
        'mt_code':       mt_code,
        'tab1': tab1,
        'tab2': tab2,
        'tab3': tab3,
        'tab4': [],
        'titles': {
            "t1": "Encoder Cable",
            "t2": "Power Cable",
            "t3": "Accessories",
            "t4": "",
        },
        'bom_list': bom_list,
    })
#-------------------------------------------------------------------
# 12. Export BOM Excel
#-------------------------------------------------------------------
def export_bom_excel(request):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    main_code = request.session.get('current_source_code')
    mt_code   = request.session.get('selected_motor')
    bom_list  = request.session.get('bom_list', [])

    if not main_code:
        return HttpResponse("Aucune configuration à exporter.")

    # ── Couleurs Yaskawa ──
    BLUE       = "0057B8"
    BLUE_LIGHT = "E8F1FB"
    TEAL       = "00A991"
    TEAL_LIGHT = "E8F7F4"
    WHITE      = "FFFFFF"
    GRAY_BG    = "F4F7FC"
    GRAY_TEXT  = "6B7A99"

    # ── Styles ──
    def cell_font(bold=False, color="000000", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def cell_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def cell_border():
        side = Side(style="thin", color="DDDDDD")
        return Border(left=side, right=side, top=side, bottom=side)

    def cell_align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    # ── Largeurs colonnes ──
    ws.column_dimensions['A'].width = 6   # #
    ws.column_dimensions['B'].width = 35  # Model Code
    ws.column_dimensions['C'].width = 20  # SAP No
    ws.column_dimensions['D'].width = 22  # Type
    ws.column_dimensions['E'].width = 40  # Description
    ws.column_dimensions['F'].width = 18  # Voltage / Power

    # ── Ligne 1 : Titre principal ──
    ws.merge_cells("A1:F1")
    ws["A1"] = "BILL OF MATERIALS — YASKAWA"
    ws["A1"].font      = cell_font(bold=True, color=WHITE, size=14)
    ws["A1"].fill      = cell_fill(BLUE)
    ws["A1"].alignment = cell_align(h="center")
    ws.row_dimensions[1].height = 36

    # ── Ligne 2 : Sous-titre (servopack + moteur) ──
    ws.merge_cells("A2:F2")
    subtitle = f"Servopack : {main_code}"
    if mt_code:
        subtitle += f"   |   Motor : {mt_code}"
    ws["A2"] = subtitle
    ws["A2"].font      = cell_font(bold=False, color=WHITE, size=10)
    ws["A2"].fill      = cell_fill("004494")
    ws["A2"].alignment = cell_align(h="center")
    ws.row_dimensions[2].height = 20

    # ── Ligne 3 : Date ──
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Generated : {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    ws["A3"].font      = cell_font(color=GRAY_TEXT, size=9)
    ws["A3"].fill      = cell_fill(GRAY_BG)
    ws["A3"].alignment = cell_align(h="right")
    ws.row_dimensions[3].height = 16

    # ── Ligne 4 : vide ──
    ws.row_dimensions[4].height = 8

    # ── Ligne 5 : En-têtes tableau ──
    headers = ["#", "Model Code", "SAP No", "Product Type", "Description", "Spec"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font      = cell_font(bold=True, color=WHITE, size=10)
        cell.fill      = cell_fill(BLUE)
        cell.alignment = cell_align(h="center")
        cell.border    = cell_border()
    ws.row_dimensions[5].height = 22

    # ── Données ──
    all_codes = [main_code]
    if mt_code:
        all_codes.append(mt_code)
    for code in bom_list:
        if code not in all_codes:
            all_codes.append(code)

    products = {p.model_code: p for p in Product.objects.filter(model_code__in=all_codes)}

    for idx, code in enumerate(all_codes, 1):
        prod = products.get(code)
        if not prod:
            continue

        row = 5 + idx

        # Couleur de fond selon le type
        is_sgd   = code.upper().startswith('SGD')
        is_sgm   = code.upper().startswith('SGM')
        bg_color = BLUE_LIGHT if is_sgd else (TEAL_LIGHT if is_sgm else WHITE)
        if idx % 2 == 0 and not is_sgd and not is_sgm:
            bg_color = GRAY_BG

        spec_str = ""
        try:
            spec_str = f"{prod.spec.voltage} / {prod.spec.power_kw} kW" if prod.spec else ""
        except Exception:
            pass

        row_data = [
            idx,
            prod.model_code,
            prod.numero_SAP or "",
            prod.product_type or "",
            prod.specification or "",
            spec_str,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = cell_fill(bg_color)
            cell.border    = cell_border()
            cell.alignment = cell_align(wrap=(col == 5))
            if col == 2:
                cell.font = cell_font(bold=True, color=BLUE if is_sgd else (TEAL if is_sgm else "000000"), size=10)
            elif col == 1:
                cell.font = cell_font(bold=True, color=GRAY_TEXT, size=9)
                cell.alignment = cell_align(h="center")
            else:
                cell.font = cell_font(size=10)

        ws.row_dimensions[row].height = 18

    # ── Ligne finale : total ──
    total_row = 5 + len(all_codes) + 1
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"] = f"Total : {len(all_codes)} item(s)"
    ws[f"A{total_row}"].font      = cell_font(bold=True, color=BLUE, size=10)
    ws[f"A{total_row}"].fill      = cell_fill(BLUE_LIGHT)
    ws[f"A{total_row}"].alignment = cell_align(h="right")
    ws.row_dimensions[total_row].height = 20

    # ── Freeze header ──
    ws.freeze_panes = "A6"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="BOM_{main_code}.xlsx"'
    wb.save(response)
    return response
#-------------------------------------------------------------------
# 13. ADD / REMOVE / CLEAR BOM
#-------------------------------------------------------------------
def add_to_solution(request, model_code, source_code):
    bom_list = list(request.session.get('bom_list', []))
    if model_code not in bom_list:
        bom_list.append(model_code)
        request.session['bom_list'] = bom_list
        request.session.modified    = True
    mt_code = request.session.get('selected_motor')
    if mt_code:
        return redirect('my_solution', sp_code=source_code, mt_code=mt_code)
    return redirect('my_solution_single', sp_code=source_code)

def remove_from_solution(request, model_code, source_code):
    bom_list = list(request.session.get('bom_list', []))
    if model_code in bom_list:
        bom_list.remove(model_code)
        request.session['bom_list'] = bom_list
        request.session.modified    = True
    mt_code = request.session.get('selected_motor')
    if mt_code:
        return redirect('my_solution', sp_code=source_code, mt_code=mt_code)
    return redirect('my_solution_single', sp_code=source_code)

def clear_solution(request):
    request.session.pop('bom_list', None)
    request.session.pop('current_source_code', None)
    return redirect('home')
#-------------------------------------------------------------------
# images 
#-------------------------------------------------------------------

#-------------------------------------------------------------------
# 14. Déconnexion
#-------------------------------------------------------------------
def logout_view(request):
    logout(request)
    return redirect('login')